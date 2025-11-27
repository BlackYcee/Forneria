"""
Views para Analytics API
Expone endpoints REST para consumir métricas desde el frontend
"""
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime, timedelta
from django.utils import timezone
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
import csv
import json
from io import BytesIO

from .services import FinanzasMetrics
from .serializers import (
    ResumenPeriodoSerializer,
    VentasDiariasSerializer,
    ProductoTopSerializer,
    VentasPorCategoriaSerializer,
    VentasPorCanalSerializer,
    ClienteTopSerializer,
    KPIsHoySerializer
)


def parse_fecha(fecha_str):
    """Helper para parsear fechas desde query params"""
    try:
        return datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None


@api_view(['GET'])
def resumen_financiero(request):
    """
    GET /analytics/finanzas/resumen/
    Query params:
        - fecha_inicio: YYYY-MM-DD (opcional, default: hace 30 días)
        - fecha_fin: YYYY-MM-DD (opcional, default: hoy)

    Retorna resumen general de ventas del periodo
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.resumen_periodo(fecha_inicio, fecha_fin)
    serializer = ResumenPeriodoSerializer(data)

    return Response(serializer.data)


@api_view(['GET'])
def ventas_diarias(request):
    """
    GET /analytics/finanzas/ventas-diarias/
    Query params:
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)

    Retorna ventas agrupadas por día (formato Chart.js)
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.ventas_diarias(fecha_inicio, fecha_fin)
    serializer = VentasDiariasSerializer(data)

    return Response(serializer.data)


@api_view(['GET'])
def ventas_por_hora(request):
    """
    GET /analytics/finanzas/ventas-por-hora/

    Retorna distribución de ventas por hora del día (últimos 7 días)
    """
    data = FinanzasMetrics.ventas_por_hora()
    serializer = VentasDiariasSerializer(data)

    return Response(serializer.data)


@api_view(['GET'])
def productos_top(request):
    """
    GET /analytics/finanzas/productos-top/
    Query params:
        - limite: int (opcional, default: 10)
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)

    Retorna top productos más vendidos
    """
    limite = int(request.GET.get('limite', 10))
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.productos_top(limite, fecha_inicio, fecha_fin)
    serializer = ProductoTopSerializer(data, many=True)

    return Response(serializer.data)


@api_view(['GET'])
def ventas_por_categoria(request):
    """
    GET /analytics/finanzas/ventas-por-categoria/
    Query params:
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)

    Retorna distribución de ventas por categoría (para gráfico de dona)
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.ventas_por_categoria(fecha_inicio, fecha_fin)
    serializer = VentasPorCategoriaSerializer(data)

    return Response(serializer.data)


@api_view(['GET'])
def ventas_por_canal(request):
    """
    GET /analytics/finanzas/ventas-por-canal/
    Query params:
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)

    Retorna ventas por canal (presencial vs delivery)
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.ventas_por_canal(fecha_inicio, fecha_fin)
    serializer = VentasPorCanalSerializer(data, many=True)

    return Response(serializer.data)


@api_view(['GET'])
def comparativa_mensual(request):
    """
    GET /analytics/finanzas/comparativa-mensual/
    Query params:
        - meses: int (opcional, default: 6)

    Retorna comparativa de ventas mensuales
    """
    meses = int(request.GET.get('meses', 6))

    data = FinanzasMetrics.comparativa_mensual(meses)
    serializer = VentasDiariasSerializer(data)

    return Response(serializer.data)


@api_view(['GET'])
def clientes_top(request):
    """
    GET /analytics/finanzas/clientes-top/
    Query params:
        - limite: int (opcional, default: 10)
        - fecha_inicio: YYYY-MM-DD (opcional)
        - fecha_fin: YYYY-MM-DD (opcional)

    Retorna top clientes por volumen de compras
    """
    limite = int(request.GET.get('limite', 10))
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.clientes_top(limite, fecha_inicio, fecha_fin)
    serializer = ClienteTopSerializer(data, many=True)

    return Response(serializer.data)


@api_view(['GET'])
def kpis_hoy(request):
    """
    GET /analytics/finanzas/kpis-hoy/

    Retorna KPIs del día actual (para cards del dashboard)
    """
    data = FinanzasMetrics.kpis_hoy()
    serializer = KPIsHoySerializer(data)

    return Response(serializer.data)


# === NUEVOS ENDPOINTS PARA MÉTRICAS AVANZADAS ===

@api_view(['GET'])
def metricas_avanzadas(request):
    """
    GET /analytics/finanzas/metricas-avanzadas/
    Métricas financieras avanzadas: ventas brutas, netas, margen descuento, etc.
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.metricas_avanzadas(fecha_inicio, fecha_fin)
    return Response(data)


@api_view(['GET'])
def ticket_promedio_segmentado(request):
    """
    GET /analytics/finanzas/ticket-segmentado/
    Ticket promedio por canal y día de semana
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.ticket_promedio_segmentado(fecha_inicio, fecha_fin)
    return Response(data)


@api_view(['GET'])
def ventas_por_dia_semana(request):
    """
    GET /analytics/finanzas/ventas-dia-semana/
    Distribución de ventas por día de la semana
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.ventas_por_dia_semana(fecha_inicio, fecha_fin)
    return Response(data)


@api_view(['GET'])
def clientes_nuevos_vs_recurrentes(request):
    """
    GET /analytics/finanzas/clientes-nuevos-recurrentes/
    Análisis de clientes nuevos vs recurrentes
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.clientes_nuevos_vs_recurrentes(fecha_inicio, fecha_fin)
    return Response(data)


@api_view(['GET'])
def heatmap_ventas(request):
    """
    GET /analytics/finanzas/heatmap-ventas/
    Heatmap de ventas por hora y día de semana
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.heatmap_ventas_hora_dia(fecha_inicio, fecha_fin)
    return Response(data)


@api_view(['GET'])
def proyeccion_ventas(request):
    """
    GET /analytics/finanzas/proyeccion/
    Proyección de ventas basada en histórico
    """
    dias = int(request.GET.get('dias', 7))
    data = FinanzasMetrics.proyeccion_ventas(dias)
    return Response(data)


@api_view(['GET'])
def comparativa_mom(request):
    """
    GET /analytics/finanzas/mom/
    Comparativa Month-over-Month con variación porcentual
    """
    meses = int(request.GET.get('meses', 6))
    data = FinanzasMetrics.comparativa_mom(meses)
    return Response(data)


@api_view(['GET'])
def alertas_automaticas(request):
    """
    GET /analytics/finanzas/alertas/
    Sistema de alertas automáticas
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    data = FinanzasMetrics.alertas_automaticas(fecha_inicio, fecha_fin)
    return Response(data)


# === FUNCIONES DE EXPORTACIÓN ===

def exportar_excel(request):
    """
    GET /analytics/finanzas/exportar/excel/
    Exporta resumen financiero completo a Excel
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    # Obtener todas las métricas
    resumen = FinanzasMetrics.resumen_periodo(fecha_inicio, fecha_fin)
    utilidad_neta = FinanzasMetrics.utilidad_neta(fecha_inicio, fecha_fin)
    roi = FinanzasMetrics.roi(fecha_inicio, fecha_fin)
    punto_equilibrio = FinanzasMetrics.punto_equilibrio(fecha_inicio, fecha_fin)
    gastos = FinanzasMetrics.gastos_operativos(fecha_inicio, fecha_fin)
    productos_rentables = FinanzasMetrics.productos_rentables(limite=20, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Análisis Financiero"

        # Encabezados
        headers_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=14)

        # === TÍTULO ===
        ws['A1'] = 'DASHBOARD FINANCIERO - FORNERIA'
        ws['A1'].font = title_font
        ws['A2'] = f"Periodo: {resumen['fecha_inicio']} a {resumen['fecha_fin']}"

        # === MÉTRICAS FINANCIERAS CLAVE ===
        row = 4
        ws[f'A{row}'] = 'MÉTRICAS FINANCIERAS CLAVE'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        ws[f'A{row}'] = 'Métrica'
        ws[f'B{row}'] = 'Valor'
        ws[f'A{row}'].fill = headers_fill
        ws[f'B{row}'].fill = headers_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font

        row += 1
        ws[f'A{row}'] = 'Ventas Totales (sin IVA)'
        ws[f'B{row}'] = f"${utilidad_neta['ventas_totales']:,.0f}"
        row += 1
        ws[f'A{row}'] = 'Costo de Ventas'
        ws[f'B{row}'] = f"${utilidad_neta['costo_ventas']:,.0f}"
        row += 1
        ws[f'A{row}'] = 'Utilidad Bruta'
        ws[f'B{row}'] = f"${utilidad_neta['utilidad_bruta']:,.0f}"
        row += 1
        ws[f'A{row}'] = 'Gastos Operativos'
        ws[f'B{row}'] = f"${utilidad_neta['gastos_operativos']:,.0f}"
        row += 1
        ws[f'A{row}'] = 'UTILIDAD NETA'
        ws[f'B{row}'] = f"${utilidad_neta['utilidad_neta']:,.0f}"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = 'Margen Utilidad Neta %'
        ws[f'B{row}'] = f"{utilidad_neta['margen_utilidad_neta_pct']}%"
        row += 1
        ws[f'A{row}'] = 'ROI %'
        ws[f'B{row}'] = f"{roi['roi_porcentaje']}%"

        # === PUNTO DE EQUILIBRIO ===
        row += 3
        ws[f'A{row}'] = 'PUNTO DE EQUILIBRIO'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        ws[f'A{row}'] = 'Transacciones Necesarias'
        ws[f'B{row}'] = f"{punto_equilibrio['transacciones_necesarias']:.0f}"
        row += 1
        ws[f'A{row}'] = 'Transacciones Actuales'
        ws[f'B{row}'] = f"{punto_equilibrio['transacciones_actuales']}"
        row += 1
        ws[f'A{row}'] = 'Porcentaje Alcanzado'
        ws[f'B{row}'] = f"{punto_equilibrio['porcentaje_alcanzado']}%"

        # === GASTOS OPERATIVOS ===
        row += 3
        ws[f'A{row}'] = 'GASTOS OPERATIVOS POR TIPO'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        ws[f'A{row}'] = 'Tipo de Gasto'
        ws[f'B{row}'] = 'Monto'
        ws[f'A{row}'].fill = headers_fill
        ws[f'B{row}'].fill = headers_fill
        ws[f'A{row}'].font = header_font
        ws[f'B{row}'].font = header_font

        for gasto in gastos['desglose']:
            row += 1
            ws[f'A{row}'] = gasto['tipo'].title()
            ws[f'B{row}'] = f"${gasto['monto']:,.0f}"

        # === PRODUCTOS MÁS RENTABLES ===
        row += 3
        ws[f'A{row}'] = 'TOP 20 PRODUCTOS MÁS RENTABLES'
        ws[f'A{row}'].font = Font(bold=True, size=12)

        row += 1
        ws[f'A{row}'] = 'Producto'
        ws[f'B{row}'] = 'Categoría'
        ws[f'C{row}'] = 'Cantidad'
        ws[f'D{row}'] = 'Ingresos'
        ws[f'E{row}'] = 'Costos'
        ws[f'F{row}'] = 'Utilidad'
        ws[f'G{row}'] = 'Margen %'

        for cell in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{cell}{row}'].fill = headers_fill
            ws[f'{cell}{row}'].font = header_font

        for p in productos_rentables:
            row += 1
            ws[f'A{row}'] = p['nombre']
            ws[f'B{row}'] = p['categoria']
            ws[f'C{row}'] = p['cantidad_vendida']
            ws[f'D{row}'] = f"${p['ingresos_totales']:,.0f}"
            ws[f'E{row}'] = f"${p['costo_total']:,.0f}"
            ws[f'F{row}'] = f"${p['utilidad_bruta']:,.0f}"
            ws[f'G{row}'] = f"{p['margen_bruto_pct']}%"

        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_financiero_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response

    except ImportError:
        # Fallback a CSV si openpyxl no está disponible
        return exportar_csv(request)


def exportar_csv(request):
    """
    GET /analytics/finanzas/exportar/csv/
    Exporta análisis financiero completo a CSV
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    # Obtener métricas
    resumen = FinanzasMetrics.resumen_periodo(fecha_inicio, fecha_fin)
    utilidad_neta = FinanzasMetrics.utilidad_neta(fecha_inicio, fecha_fin)
    roi = FinanzasMetrics.roi(fecha_inicio, fecha_fin)
    punto_equilibrio = FinanzasMetrics.punto_equilibrio(fecha_inicio, fecha_fin)
    productos_rentables = FinanzasMetrics.productos_rentables(limite=20, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    gastos = FinanzasMetrics.gastos_operativos(fecha_inicio, fecha_fin)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=reporte_financiero_{timezone.now().strftime("%Y%m%d")}.csv'

    writer = csv.writer(response)

    # Título
    writer.writerow(['DASHBOARD FINANCIERO - FORNERIA'])
    writer.writerow([f"Periodo: {resumen['fecha_inicio']} a {resumen['fecha_fin']}"])
    writer.writerow([])

    # Métricas Financieras Clave
    writer.writerow(['MÉTRICAS FINANCIERAS CLAVE'])
    writer.writerow(['Métrica', 'Valor'])
    writer.writerow(['Ventas Totales (sin IVA)', f"${utilidad_neta['ventas_totales']:,.0f}"])
    writer.writerow(['Costo de Ventas', f"${utilidad_neta['costo_ventas']:,.0f}"])
    writer.writerow(['Utilidad Bruta', f"${utilidad_neta['utilidad_bruta']:,.0f}"])
    writer.writerow(['Gastos Operativos', f"${utilidad_neta['gastos_operativos']:,.0f}"])
    writer.writerow(['UTILIDAD NETA', f"${utilidad_neta['utilidad_neta']:,.0f}"])
    writer.writerow(['Margen Utilidad Neta %', f"{utilidad_neta['margen_utilidad_neta_pct']}%"])
    writer.writerow(['ROI %', f"{roi['roi_porcentaje']}%"])
    writer.writerow([])

    # Punto de Equilibrio
    writer.writerow(['PUNTO DE EQUILIBRIO'])
    writer.writerow(['Métrica', 'Valor'])
    writer.writerow(['Transacciones Necesarias', f"{punto_equilibrio['transacciones_necesarias']:.0f}"])
    writer.writerow(['Transacciones Actuales', f"{punto_equilibrio['transacciones_actuales']}"])
    writer.writerow(['Porcentaje Alcanzado', f"{punto_equilibrio['porcentaje_alcanzado']}%"])
    writer.writerow([])

    # Gastos Operativos
    writer.writerow(['GASTOS OPERATIVOS POR TIPO'])
    writer.writerow(['Tipo de Gasto', 'Monto'])
    for gasto in gastos['desglose']:
        writer.writerow([gasto['tipo'].title(), f"${gasto['monto']:,.0f}"])
    writer.writerow([])

    # Top Productos Rentables
    writer.writerow(['TOP 20 PRODUCTOS MÁS RENTABLES'])
    writer.writerow(['Producto', 'Categoría', 'Cantidad', 'Ingresos', 'Costos', 'Utilidad', 'Margen %'])
    for p in productos_rentables:
        writer.writerow([
            p['nombre'],
            p['categoria'],
            p['cantidad_vendida'],
            f"${p['ingresos_totales']:,.0f}",
            f"${p['costo_total']:,.0f}",
            f"${p['utilidad_bruta']:,.0f}",
            f"{p['margen_bruto_pct']}%"
        ])

    return response


def exportar_pdf(request):
    """
    GET /analytics/finanzas/exportar/pdf/
    Exporta análisis financiero completo a PDF
    """
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch

        # Obtener métricas
        resumen = FinanzasMetrics.resumen_periodo(fecha_inicio, fecha_fin)
        utilidad_neta = FinanzasMetrics.utilidad_neta(fecha_inicio, fecha_fin)
        roi = FinanzasMetrics.roi(fecha_inicio, fecha_fin)
        punto_equilibrio = FinanzasMetrics.punto_equilibrio(fecha_inicio, fecha_fin)
        productos_rentables = FinanzasMetrics.productos_rentables(limite=10, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
        )
        elements.append(Paragraph("Dashboard Financiero - Forneria", title_style))
        elements.append(Paragraph(f"Periodo: {resumen['fecha_inicio']} a {resumen['fecha_fin']}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        # Tabla Métricas Financieras
        elements.append(Paragraph("Métricas Financieras Clave", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        data_finanzas = [
            ['Métrica', 'Valor'],
            ['Ventas Totales (sin IVA)', f"${utilidad_neta['ventas_totales']:,.0f}"],
            ['Costo de Ventas', f"${utilidad_neta['costo_ventas']:,.0f}"],
            ['Utilidad Bruta', f"${utilidad_neta['utilidad_bruta']:,.0f}"],
            ['Gastos Operativos', f"${utilidad_neta['gastos_operativos']:,.0f}"],
            ['UTILIDAD NETA', f"${utilidad_neta['utilidad_neta']:,.0f}"],
            ['Margen Utilidad Neta %', f"{utilidad_neta['margen_utilidad_neta_pct']}%"],
            ['ROI %', f"{roi['roi_porcentaje']}%"],
        ]

        table_finanzas = Table(data_finanzas, colWidths=[3*inch, 2*inch])
        table_finanzas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            # Resaltar Utilidad Neta
            ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#E8F4F8')),
        ]))

        elements.append(table_finanzas)
        elements.append(Spacer(1, 0.5*inch))

        # Punto de Equilibrio
        elements.append(Paragraph("Punto de Equilibrio", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        data_equilibrio = [
            ['Métrica', 'Valor'],
            ['Transacciones Necesarias', f"{punto_equilibrio['transacciones_necesarias']:.0f}"],
            ['Transacciones Actuales', f"{punto_equilibrio['transacciones_actuales']}"],
            ['Porcentaje Alcanzado', f"{punto_equilibrio['porcentaje_alcanzado']}%"],
        ]

        table_equilibrio = Table(data_equilibrio, colWidths=[3*inch, 2*inch])
        table_equilibrio.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))

        elements.append(table_equilibrio)
        elements.append(Spacer(1, 0.5*inch))

        # Top Productos Rentables
        elements.append(Paragraph("Top 10 Productos Más Rentables", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        data_productos = [['Producto', 'Cantidad', 'Ingresos', 'Utilidad', 'Margen%']]
        for p in productos_rentables:
            data_productos.append([
                p['nombre'][:25],
                p['cantidad_vendida'],
                f"${p['ingresos_totales']:,.0f}",
                f"${p['utilidad_bruta']:,.0f}",
                f"{p['margen_bruto_pct']}%"
            ])

        table_productos = Table(data_productos, colWidths=[2.2*inch, 0.8*inch, 1.1*inch, 1.1*inch, 0.8*inch])
        table_productos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))

        elements.append(table_productos)

        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=reporte_financiero_{timezone.now().strftime("%Y%m%d")}.pdf'
        return response

    except ImportError:
        return Response({
            'error': 'La librería reportlab no está instalada. Instálala con: pip install reportlab'
        }, status=500)


# Vista HTML del Dashboard
def dashboard_finanzas(request):
    """
    Vista HTML del dashboard financiero mejorado
    Renderiza template con todas las métricas avanzadas cargadas
    """
    # Parsear filtros de fecha desde request
    fecha_inicio = parse_fecha(request.GET.get('fecha_inicio'))
    fecha_fin = parse_fecha(request.GET.get('fecha_fin'))

    # Obtener todas las métricas
    kpis = FinanzasMetrics.kpis_hoy()
    resumen = FinanzasMetrics.resumen_periodo(fecha_inicio, fecha_fin)
    metricas_avanzadas = FinanzasMetrics.metricas_avanzadas(fecha_inicio, fecha_fin)
    productos_top = FinanzasMetrics.productos_top(limite=10, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    categorias = FinanzasMetrics.ventas_por_categoria(fecha_inicio, fecha_fin)
    ventas_canal = FinanzasMetrics.ventas_por_canal(fecha_inicio, fecha_fin)
    clientes_top = FinanzasMetrics.clientes_top(limite=5, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    ventas_diarias = FinanzasMetrics.ventas_diarias(fecha_inicio, fecha_fin)
    ventas_hora = FinanzasMetrics.ventas_por_hora()
    comparativa_mensual = FinanzasMetrics.comparativa_mensual(6)
    ticket_segmentado = FinanzasMetrics.ticket_promedio_segmentado(fecha_inicio, fecha_fin)
    ventas_dia_semana = FinanzasMetrics.ventas_por_dia_semana(fecha_inicio, fecha_fin)
    clientes_nuevos = FinanzasMetrics.clientes_nuevos_vs_recurrentes(fecha_inicio, fecha_fin)
    heatmap = FinanzasMetrics.heatmap_ventas_hora_dia(fecha_inicio, fecha_fin)
    proyeccion = FinanzasMetrics.proyeccion_ventas(7)
    mom = FinanzasMetrics.comparativa_mom(6)
    alertas = FinanzasMetrics.alertas_automaticas()

    # NUEVAS MÉTRICAS FINANCIERAS
    utilidad_bruta = FinanzasMetrics.utilidad_bruta(fecha_inicio, fecha_fin)
    gastos_operativos = FinanzasMetrics.gastos_operativos(fecha_inicio, fecha_fin)
    utilidad_neta = FinanzasMetrics.utilidad_neta(fecha_inicio, fecha_fin)
    roi = FinanzasMetrics.roi(fecha_inicio, fecha_fin)
    punto_equilibrio = FinanzasMetrics.punto_equilibrio(fecha_inicio, fecha_fin)
    productos_rentables = FinanzasMetrics.productos_rentables(limite=10, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    flujo_caja = FinanzasMetrics.flujo_caja(fecha_inicio, fecha_fin)

    # Helper function to convert Decimal to float for JSON serialization
    def clean_for_json(obj):
        """Recursively convert Decimal to float in dict/list structures"""
        if isinstance(obj, dict):
            return {k: clean_for_json(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [clean_for_json(item) for item in obj]
        elif hasattr(obj, '__float__'):
            return float(obj)
        else:
            return obj

    context = {
        'kpis': kpis,
        'resumen': resumen,
        'metricas_avanzadas': metricas_avanzadas,
        'productos_top': productos_top,
        'categorias': categorias,
        'ventas_canal': ventas_canal,
        'clientes_top': clientes_top,
        'ventas_diarias': ventas_diarias,
        'ventas_hora': ventas_hora,
        'comparativa_mensual': comparativa_mensual,
        'ticket_segmentado': ticket_segmentado,
        'ventas_dia_semana': ventas_dia_semana,
        'clientes_nuevos': clientes_nuevos,
        'heatmap': heatmap,
        'proyeccion': proyeccion,
        'mom': mom,
        'alertas': alertas,

        # NUEVAS MÉTRICAS FINANCIERAS
        'utilidad_bruta': utilidad_bruta,
        'gastos_operativos': gastos_operativos,
        'utilidad_neta': utilidad_neta,
        'roi': roi,
        'punto_equilibrio': punto_equilibrio,
        'productos_rentables': productos_rentables,
        'flujo_caja': flujo_caja,

        'fecha_actual': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
        'fecha_inicio': fecha_inicio.isoformat() if fecha_inicio else '',
        'fecha_fin': fecha_fin.isoformat() if fecha_fin else '',

        # JSON-encoded versions for JavaScript (safe from template syntax errors)
        'ventas_diarias_json': json.dumps(clean_for_json(ventas_diarias)),
        'proyeccion_json': json.dumps(clean_for_json(proyeccion)),
        'mom_json': json.dumps(clean_for_json(mom)),
        'ventas_dia_semana_json': json.dumps(clean_for_json(ventas_dia_semana)),
        'ventas_canal_json': json.dumps(clean_for_json(ventas_canal)),
        'categorias_json': json.dumps(clean_for_json(categorias)),
        'clientes_nuevos_json': json.dumps(clean_for_json(clientes_nuevos)),
        'ventas_hora_json': json.dumps(clean_for_json(ventas_hora)),
        'flujo_caja_json': json.dumps(clean_for_json(flujo_caja)),
        'gastos_desglose_json': json.dumps(clean_for_json(gastos_operativos['desglose'])),
    }

    return render(request, 'dashboard_finanzas.html', context)
